import streamlit as st
import pandas as pd
import os
import signal


if "poly_ready" not in st.session_state:
    st.session_state.poly_ready = False

if "show_chr_map" not in st.session_state:
    st.session_state.show_chr_map = False

if "parent_uploaded" not in st.session_state:
    st.session_state.parent_uploaded = False

if "marker_uploaded" not in st.session_state:
    st.session_state.marker_uploaded = False

from db import init_db
from io_utils import upload_bc_matrix
from analysis import analyze_bc

# -------------------------------------------------
# Initialize
# -------------------------------------------------
init_db()

st.set_page_config(
    page_title="MABS Tool",
    layout="center"
)


st.markdown(
    """
    <style>

    section.main > div {
        max-width: 1500px;
        padding-left: 2rem;
        padding-right: 2rem;
    }

    </style>
    """,
    unsafe_allow_html=True
)
# ---------------------------------------------
# UI CSS (Equal width tabs)
# ---------------------------------------------
st.markdown(
    """
    <style>

    /* Make tab bar use full width */
    div[data-baseweb="tab-list"] {
        display: flex;
        width: 100%;
        justify-content: space-between;
    }

    /* Each tab gets equal width */
    button[data-baseweb="tab"] {
        flex: 1 1 0%;
        justify-content: center;
        text-align: center;
        padding: 10px 0px !important;
    }

    /* Tab label */
    button[data-baseweb="tab"] > div {
        font-size: 18px !important;
        font-weight: 600;
        width: 100%;
        text-align: center;
    }

    </style>
    """,
    unsafe_allow_html=True
)

# -------------------------------------------------
# Header
# -------------------------------------------------
st.markdown(
    "<h1 style='text-align: center;'>🌿 MABS Data Analysis Tool</h1>",
    unsafe_allow_html=True
)



st.markdown(
    "<p style='text-align:center; font-size:14px;'>"
    "Version 1.1"
    "</p>",
    unsafe_allow_html=True
) 

st.markdown("---")

# -------------------------------------------------
# MAIN TABS
# -------------------------------------------------
tab1, tab2, tab3 = st.tabs([
    "🧬 Parent Genotyping data",
    "🔍 Get Polymorphic Markers",
    "📊 BG Recovery Analysis"
])

# =================================================
# TAB 1: PARENT GENOTYPING
# =================================================
with tab1:

    st.subheader("Upload Parent Genotyping Data")

    
    file = st.file_uploader(
        "Upload Excel file",
        type="xlsx",
        key="parent_upload_file"
    )

    if st.button("Upload Genotyping Data", key="parent_upload_btn"):

        if file is None:
            st.error("Please upload the parent genotyping file.")
            st.stop()

        parent_df = pd.read_excel(file)

        # Standardize IDs
        parent_df.iloc[:, 0] = (
            parent_df.iloc[:, 0]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        st.session_state["parent_df"] = parent_df
        st.session_state.parent_uploaded = True

        st.session_state.parent_uploaded = True
 
        st.success("Parent genotyping data uploaded successfully.")

    #Block for Marker position data upload

    st.markdown("---")
    st.subheader("Upload SNP Marker Position File")

    pos_file = st.file_uploader(
        "Upload Excel file",
        type="xlsx",
        key="marker_pos_file"
    )


    if st.button("Upload Marker Position Data"):

        if pos_file is None:
            st.error("Please upload the marker position file.")
            st.stop()

        from analysis import save_marker_position_file

        marker_sheet = pd.read_excel(pos_file, sheet_name=0)
        chrom_sheet = pd.read_excel(pos_file, sheet_name=1)

        # Standardize column names
        marker_sheet.columns = ["marker", "chr", "position_bp"]
        chrom_sheet.columns = ["chr", "chr_length_bp"]

        # Normalize
        marker_sheet["marker"] = marker_sheet["marker"].astype(str).str.strip().str.upper()
        marker_sheet["chr"] = marker_sheet["chr"].astype(str).str.strip().str.upper()

        chrom_sheet["chr"] = chrom_sheet["chr"].astype(str).str.strip().str.upper()

        st.session_state["marker_df"] = marker_sheet
        st.session_state["chrom_df"] = chrom_sheet

        st.session_state.marker_uploaded = True

        st.success(
            f"Marker position data uploaded: "
            f"{len(marker_sheet)} markers, "
            f"{len(chrom_sheet)} chromosomes"
        )
        
        
# =================================================
# TAB 2: RP–DP POLYMORPHIC MARKERS
# =================================================
with tab2:

    st.subheader("Polymorphic Marker List")
    rp = None
    dp = None
    if not st.session_state.parent_uploaded:
        st.info("Please upload Parent Genotyping Data in Tab 1.")
    
    else:
    
        parent_df = st.session_state["parent_df"]

        parent_names = (
            parent_df.columns[1:]
            .astype(str)
            .str.strip()
            .tolist()
        )
            
        rp = st.selectbox(
            "Recurrent Parent",
            options=parent_names,
            index=0,
            key="poly_rp"
        )

        dp = st.selectbox(
            "Donor Parent",
            options=parent_names,
            index=1 if len(parent_names) > 1 else 0,
            key="poly_dp"
        )

    if st.session_state.parent_uploaded and st.button(
        "View Polymorphic Markers",
        key="poly_btn"
    ):

        if rp == dp:
            st.error("Recurrent Parent and Donor Parent cannot be the same.")
            st.stop()

        from analysis import get_polymorphic_markers

        poly = get_polymorphic_markers(
            st.session_state["parent_df"],
            rp,
            dp
        )

        if poly.empty:
            st.warning("No polymorphic markers found for the selected RP/DP pair")
            st.stop()

        st.metric("🧬 Polymorphic markers identified", len(poly))
        st.dataframe(poly, use_container_width=True)

        # -------------------------------
        # Excel download
        # -------------------------------
        import io

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            poly.to_excel(
                writer,
                index=False,
                sheet_name="Polymorphic_Markers"
            )

        buffer.seek(0)

        st.download_button(
            "Download Polymorphic Markers (Excel)",
            buffer,
            file_name="Polymorphic_markers.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.markdown("---")
    
    if st.session_state.parent_uploaded and st.button(
        "View Chromosome-wise Marker Map"
    ):

        from analysis import get_marker_status_for_map, plot_chromosome_map

        marker_status = get_marker_status_for_map(
            st.session_state["parent_df"],
            rp,
            dp
        )

        fig = plot_chromosome_map(
            marker_status,
            st.session_state["marker_df"],
            st.session_state["chrom_df"],
            rp,
            dp
        )

        st.pyplot(fig)

        import io

        png = io.BytesIO()
        fig.savefig(png, format="png", dpi=300, bbox_inches="tight")
        png.seek(0)

        jpg = io.BytesIO()
        fig.savefig(jpg, format="jpg", dpi=300, bbox_inches="tight")
        jpg.seek(0)

        st.download_button(
            "Download Map (PNG)",
            png,
            file_name=f"Chr_Map_{rp}_{dp}.png",
            mime="image/png"
        )

        st.download_button(
            "Download Map (JPG)",
            jpg,
            file_name=f"Chr_Map_{rp}_{dp}.jpg",
            mime="image/jpeg"
        )


# =================================================
# TAB 3: UPLOAD BC GENOTYPING
# =================================================
with tab3:

    st.subheader("Upload BC Genotyping Data")

    
    file = st.file_uploader(
        "Upload BC Excel file",
        type="xlsx",
        key="bc_upload_file"
    )

    if file is not None:

        
        bc_df = pd.read_excel(
            file,
            keep_default_na=False
        )

        # Standardize line names
        bc_df.iloc[:, 0] = (
            bc_df.iloc[:, 0]
            .astype(str)
            .str.strip()
        )

        st.session_state["bc_df"] = bc_df

        line_names = bc_df.iloc[:, 0].tolist()

        marker_count = bc_df.shape[1] - 1

        plant_count = max(len(line_names) - 2, 0)

        st.success("BC genotyping data uploaded successfully.")

        st.info(f"""
        
        Detected rows : {len(line_names)}

        Detected markers : {marker_count}

        Detected BC plants : {plant_count}
        """
        )

# =================================================
# BG RECOVERY & RANKING
# =================================================

    # ---------------------------------------------
    # INFO PANEL (OPTION 3 – SECTION PANEL)
    # ---------------------------------------------
    

    st.subheader("BG Recovery & Ranking")

    if "bc_df" not in st.session_state:
        st.info("Upload a BC genotyping file to continue.")
        rp = None
        dp = None
        line_names = []
    else:

        bc_df = st.session_state["bc_df"]

        line_names = (
            bc_df.iloc[:, 0]
            .astype(str)
            .str.strip()
            .tolist()
        )

        rp = st.selectbox(
            "Detected Recurrent Parent",
            options=line_names,
            index=0,
            key="recovery_rp"
        )

        dp = st.selectbox(
            "Detected Donor Parent",
            options=line_names,
            index=1 if len(line_names) > 1 else 0,
            key="recovery_dp"
        )

        # Remaining BG Recovery widgets...

    
    if "bc_df" in st.session_state:

        st.caption(
            f"BC plants available: {max(len(line_names)-2, 0)}"
        )

        st.markdown("---")

        effective_markers = st.number_input(
            "Effective Markers",
            min_value=1,
            step=1,
            key="effective_markers"
        )

        polymorphic_markers = st.number_input(
            "Polymorphic Markers",
            min_value=0,
            max_value=effective_markers,
            step=1,
            key="polymorphic_markers"
        )

        monomorphic_markers = effective_markers - polymorphic_markers

        st.metric(
            "Monomorphic Markers",
            monomorphic_markers
        )    


    # ---------------------------------------------
    # RUN ANALYSIS
    # ---------------------------------------------
    if "bc_df" in st.session_state and st.button("Analyze BG Recovery", key="recovery_btn"):

        
        if "bc_df" not in st.session_state:
            st.error("Please upload BC genotyping data first")
            st.stop()
        
        if rp == dp:
            st.error("Recurrent Parent and Donor Parent cannot be the same.")
            st.stop()

        if effective_markers == 0:
            st.error("Please enter the number of effective markers.")
            st.stop()

        res = analyze_bc(
            st.session_state["bc_df"],
            rp,
            dp,
            effective_markers,
            polymorphic_markers
        )

        # -----------------------------------------
        # OPTION 2: COLORED TABLE
        # -----------------------------------------
        def highlight_top_rank(row):
            if row["Rank"] == 1:
                return ["background-color: #C8E6C9"] * len(row)
            return [""] * len(row)

        styled_res = (
            res.style
            .apply(highlight_top_rank, axis=1)
        )

        st.dataframe(styled_res, use_container_width=True)

        # -----------------------------------------
        # OPTION 4: FORMATTED EXCEL OUTPUT
        # -----------------------------------------
        from io import BytesIO
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        excel_buffer = BytesIO()
        res.to_excel(
            excel_buffer,
            index=False,
            sheet_name="BC_Recovery"
        )
        excel_buffer.seek(0)

        wb = load_workbook(excel_buffer)
        ws = wb.active

        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Auto column width
        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col
            )
            ws.column_dimensions[
                get_column_letter(col[0].column)
            ].width = max_len + 2

        # Highlight Rank = 1 rows
        green_fill = PatternFill(
            start_color="C8E6C9",
            end_color="C8E6C9",
            fill_type="solid"
        )

        rank_col_idx = None
        for i, cell in enumerate(ws[1], start=1):
            if cell.value == "Rank":
                rank_col_idx = i
                break

        if rank_col_idx:
            for row in ws.iter_rows(min_row=2):
                if row[rank_col_idx - 1].value == 1:
                    for cell in row:
                        cell.fill = green_fill

        final_buffer = BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        st.download_button(
            label="Download BG Recovery (Excel)",
            data=final_buffer,
            file_name="BC_Recovery_Ranking.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="recovery_download"
        )


# -------------------------------------------------
# Footer
# -------------------------------------------------
st.markdown(
    "<hr style='margin-top:40px;'>"
    "<center><small>Developed by SHIVAJI AJINATH LAVALE @ KSCL</small></center>",
    unsafe_allow_html=True
)


